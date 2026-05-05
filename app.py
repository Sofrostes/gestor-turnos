import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from collections import defaultdict
from io import BytesIO

# ============================================================
# CONFIGURACIÓN DE LA PÁGINA
# ============================================================
st.set_page_config(
    page_title="Gestor de Intercambios de Turnos",
    page_icon="🏭",
    layout="wide"
)

# ============================================================
# REGLAS DE VALIDACIÓN (MODIFICA AQUÍ SEGÚN NECESITES)
# ============================================================
DIAS_PROHIBIDOS = [1, 5, 8, 15, 22]
MAX_INTERCAMBIOS_POR_AGENTE = 3
TURNOS_PROHIBIDOS = ["D", "E", "VC"]
MAX_DIAS_CONSECUTIVOS = 6
MAX_DIAS_MES = 22

ZONAS_COMPATIBLES = {
    "LINEA_3": ["LINEA_3", "ALBORAYA"],
    "ZONA_6": ["ZONA_6", "ZONA_7"],
    "ZONA_7": ["ZONA_6", "ZONA_7", "ZONA_8"],
    "ZONA_8": ["ZONA_7", "ZONA_8"],
    "AEROPORT": ["AEROPORT", "MARITIM"],
    "MARITIM": ["AEROPORT", "MARITIM"],
}

# Días de la semana en español para mayo 2026
NOMBRES_DIAS = ["L", "M", "X", "J", "V", "S", "D"]
DIAS_SEMANA_MAYO = []
for dia in range(1, 32):
    indice = (dia + 3) % 7
    DIAS_SEMANA_MAYO.append(NOMBRES_DIAS[indice])


# ============================================================
# CLASE PRINCIPAL
# ============================================================

class GestorTurnosWeb:
    def __init__(self, archivo_bytes=None):
        self.wb = None
        self.ws = None
        self.agentes = {}
        self.turnos = {}
        self.intercambios = []
        self.contador_intercambios = defaultdict(int)
        self.deudas = defaultdict(int)
        
        if archivo_bytes:
            self.cargar_archivo(archivo_bytes)
    
    def cargar_archivo(self, archivo_bytes):
        self.wb = load_workbook(BytesIO(archivo_bytes))
        self.ws = self.wb["MAYO 2026"]
        self.detectar_agentes()
        self.cargar_turnos()
        return True
    
    def detectar_agentes(self):
        zonas_keywords = {
            "ALBORAYA": "ALBORAYA", "AV. DEL CID": "AV_CID", "ALAMEDA": "ALAMEDA",
            "FOIOS": "FOIOS", "MASSAMAGRELL": "MASSAMAGRELL", "AEROPORT": "AEROPORT",
            "MARITIM": "MARITIM", "TALLER": "TALLER", "LÍNEA 3": "LINEA_3",
            "ZONA 6": "ZONA_6", "ZONA 7": "ZONA_7", "ZONA 8": "ZONA_8"
        }
        
        zona_actual = "DESCONOCIDA"
        self.agentes = {}  # Limpiar antes de detectar
        
        for fila in range(1, 300):
            celda_a = self.ws[f"A{fila}"].value
            if celda_a and isinstance(celda_a, str):
                for keyword, zona in zonas_keywords.items():
                    if keyword.upper() in celda_a.upper():
                        zona_actual = zona
                        break
            
            nombre_celda = self.ws[f"C{fila}"].value
            if nombre_celda and isinstance(nombre_celda, str):
                nombre = nombre_celda.strip()
                # Filtrar nombres válidos (no cabeceras, no números, no vacíos)
                if nombre and nombre not in ["0", "AGENTE", "COD.", "EST.", "NOMBRE", "", "DESPLAZADO 1", "DESPLAZADO 2"]:
                    if not nombre.isdigit() and len(nombre) > 2:
                        # Verificar que no sea una cabecera de zona
                        es_zona = False
                        for keyword in zonas_keywords.keys():
                            if keyword.upper() in nombre.upper():
                                es_zona = True
                                break
                        if not es_zona:
                            self.agentes[nombre] = {"fila": fila, "zona": zona_actual}
    
    def cargar_turnos(self):
        for nombre, info in self.agentes.items():
            fila = info["fila"]
            for dia in range(1, 32):
                col_num = (dia * 2) + 3
                col = get_column_letter(col_num)
                celda = self.ws[f"{col}{fila}"]
                turno = celda.value if celda.value else ""
                self.turnos[(nombre, dia)] = str(turno).strip() if turno else ""
    
    def obtener_turno(self, nombre, dia):
        return self.turnos.get((nombre, dia), "")
    
    def obtener_todos_turnos(self, nombre):
        return {dia: self.obtener_turno(nombre, dia) for dia in range(1, 32)}
    
    def obtener_tipo_turno(self, turno):
        if not turno:
            return "L"
        turno_str = str(turno).upper()
        if turno_str.endswith("F"):
            return "F"
        elif turno_str.endswith("S"):
            return "S"
        elif turno_str.endswith("N"):
            return "N"
        return "L"
    
    def turnos_compatibles(self, turno1, turno2):
        if not turno1 or not turno2:
            return True
        if turno1 in TURNOS_PROHIBIDOS or turno2 in TURNOS_PROHIBIDOS:
            return False
        return self.obtener_tipo_turno(turno1) == self.obtener_tipo_turno(turno2)
    
    def contar_dias_trabajados(self, nombre):
        trabajados = 0
        for dia in range(1, 32):
            turno = self.obtener_turno(nombre, dia)
            if turno and turno not in ["D", "E", ""]:
                trabajados += 1
        return trabajados
    
    def contar_consecutivos(self, nombre, dia_cambio):
        trabajados = []
        for dia in range(1, 32):
            turno = self.obtener_turno(nombre, dia)
            trabajados.append(turno and turno not in ["D", "E", ""])
        
        idx = dia_cambio - 1
        
        back = 0
        for i in range(idx - 1, -1, -1):
            if trabajados[i]:
                back += 1
            else:
                break
        
        forward = 0
        for i in range(idx + 1, 31):
            if trabajados[i]:
                forward += 1
            else:
                break
        
        return back + forward + 1
    
    def validar_intercambio(self, nombre1, nombre2, dia):
        errores = []
        
        if self.contador_intercambios[nombre1] >= MAX_INTERCAMBIOS_POR_AGENTE:
            errores.append(f"{nombre1} ha alcanzado el máximo de {MAX_INTERCAMBIOS_POR_AGENTE} intercambios")
        if self.contador_intercambios[nombre2] >= MAX_INTERCAMBIOS_POR_AGENTE:
            errores.append(f"{nombre2} ha alcanzado el máximo de {MAX_INTERCAMBIOS_POR_AGENTE} intercambios")
        
        zona1 = self.agentes[nombre1]["zona"]
        zona2 = self.agentes[nombre2]["zona"]
        if zona1 != zona2:
            compatible = False
            for base, aceptadas in ZONAS_COMPATIBLES.items():
                if zona1 in [base] + aceptadas and zona2 in [base] + aceptadas:
                    compatible = True
                    break
            if not compatible:
                errores.append(f"Zonas incompatibles: {zona1} vs {zona2}")
        
        turno1 = self.obtener_turno(nombre1, dia)
        turno2 = self.obtener_turno(nombre2, dia)
        
        if not self.turnos_compatibles(turno1, turno2):
            errores.append(f"Turnos incompatibles: '{turno1}' vs '{turno2}'")
        
        orig1 = self.turnos.get((nombre1, dia))
        orig2 = self.turnos.get((nombre2, dia))
        
        self.turnos[(nombre1, dia)] = turno2
        self.turnos[(nombre2, dia)] = turno1
        
        if self.contar_dias_trabajados(nombre1) > MAX_DIAS_MES:
            errores.append(f"{nombre1} excedería los {MAX_DIAS_MES} días trabajados")
        if self.contar_dias_trabajados(nombre2) > MAX_DIAS_MES:
            errores.append(f"{nombre2} excedería los {MAX_DIAS_MES} días trabajados")
        
        if self.contar_consecutivos(nombre1, dia) > MAX_DIAS_CONSECUTIVOS:
            errores.append(f"{nombre1} excedería los {MAX_DIAS_CONSECUTIVOS} días consecutivos")
        if self.contar_consecutivos(nombre2, dia) > MAX_DIAS_CONSECUTIVOS:
            errores.append(f"{nombre2} excedería los {MAX_DIAS_CONSECUTIVOS} días consecutivos")
        
        self.turnos[(nombre1, dia)] = orig1
        self.turnos[(nombre2, dia)] = orig2
        
        return len(errores) == 0, errores
    
    def ejecutar_intercambio(self, nombre1, nombre2, dia, es_descanso=False):
        valido, errores = self.validar_intercambio(nombre1, nombre2, dia)
        if not valido:
            return False, errores
        
        turno1 = self.obtener_turno(nombre1, dia)
        turno2 = self.obtener_turno(nombre2, dia)
        
        fila1 = self.agentes[nombre1]["fila"]
        fila2 = self.agentes[nombre2]["fila"]
        col_num = (dia * 2) + 3
        col = get_column_letter(col_num)
        
        if es_descanso:
            self.ws[f"{col}{fila1}"].value = "D"
            self.ws[f"{col}{fila2}"].value = turno1
            self.deudas[(nombre1, nombre2)] += 1
        else:
            self.ws[f"{col}{fila1}"].value = turno2
            self.ws[f"{col}{fila2}"].value = turno1
        
        self.turnos[(nombre1, dia)] = "D" if es_descanso else turno2
        self.turnos[(nombre2, dia)] = turno1 if es_descanso else turno1
        
        self.intercambios.append({
            "nombre1": nombre1, "nombre2": nombre2, "dia": dia,
            "turno1": turno1, "turno2": turno2, "es_descanso": es_descanso
        })
        self.contador_intercambios[nombre1] += 1
        self.contador_intercambios[nombre2] += 1
        
        return True, ["✅ Intercambio realizado correctamente"]
    
    def obtener_bytes(self):
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output


# ============================================================
# FUNCIONES DE VISUALIZACIÓN
# ============================================================

def mostrar_cuadrante_mensual(turnos, nombre_agente, gestor):
    """Muestra una tabla con todos los días del mes y sus turnos"""
    
    data = []
    for dia in range(1, 32):
        turno = turnos.get(dia, "")
        dia_semana = DIAS_SEMANA_MAYO[dia-1]
        
        # Color según tipo de día
        if turno in ["D", "E"]:
            icono = "🟢"
        elif turno in ["VC", "LC", "M", "MOD"]:
            icono = "🟡"
        elif turno and turno not in ["", None]:
            icono = "🔵"
        else:
            icono = "⚪"
        
        data.append({
            "Día": dia,
            "Semana": dia_semana,
            "Turno": turno if turno else "-",
            "": icono
        })
    
    df = pd.DataFrame(data)
    
    # Estadísticas
    trabajados = sum(1 for d in range(1, 32) if turnos.get(d) and turnos.get(d) not in ["D", "E", ""])
    descansos = sum(1 for d in range(1, 32) if turnos.get(d) in ["D", "E"])
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("📊 Días trabajados", trabajados, delta=f"máx {MAX_DIAS_MES}")
    with col2:
        st.metric("🟢 Descansos", descansos)
    with col3:
        st.metric("🔄 Intercambios", gestor.contador_intercambios.get(nombre_agente, 0))
    
    st.dataframe(df, use_container_width=True, hide_index=True)
    
    with st.expander("📖 Leyenda"):
        st.markdown("""
        - 🔵 **Azul**: Día trabajado
        - 🟢 **Verde**: Descanso (D/E)
        - 🟡 **Amarillo**: Servicios especiales
        - ⚪ **Blanco**: Sin asignar
        """)


# ============================================================
# INTERFAZ DE USUARIO
# ============================================================

st.title("🏭 Gestor de Intercambios de Turnos - Mayo 2026")
st.markdown("---")

if "gestor" not in st.session_state:
    st.session_state.gestor = None
if "archivo_cargado" not in st.session_state:
    st.session_state.archivo_cargado = False

# Sidebar
with st.sidebar:
    st.header("📁 Cargar archivo")
    
    uploaded_file = st.file_uploader(
        "Selecciona el archivo turnos_mayo_2026.xlsx",
        type=["xlsx"],
        help="Sube el archivo Excel con los turnos de Mayo 2026"
    )
    
    if uploaded_file and not st.session_state.archivo_cargado:
        with st.spinner("Cargando archivo..."):
            gestor = GestorTurnosWeb()
            if gestor.cargar_archivo(uploaded_file.read()):
                st.session_state.gestor = gestor
                st.session_state.archivo_cargado = True
                st.success(f"✅ Cargados {len(gestor.agentes)} agentes")
                
                # Mostrar los primeros agentes para depuración
                with st.expander("📋 Ver agentes detectados"):
                    for i, nombre in enumerate(list(gestor.agentes.keys())[:20]):
                        st.write(f"{i+1}. {nombre}")
                    if len(gestor.agentes) > 20:
                        st.write(f"... y {len(gestor.agentes)-20} más")
            else:
                st.error("❌ Error al cargar el archivo")
    
    if st.session_state.archivo_cargado:
        st.markdown("---")
        st.header("📋 Reglas activas")
        st.info(f"""
        - Días prohibidos: {DIAS_PROHIBIDOS}
        - Máx intercambios/agente: {MAX_INTERCAMBIOS_POR_AGENTE}
        - Turnos prohibidos: {TURNOS_PROHIBIDOS}
        - Máx días consecutivos: {MAX_DIAS_CONSECUTIVOS}
        - Máx días/mes: {MAX_DIAS_MES}
        """)

# Contenido principal
if not st.session_state.archivo_cargado:
    st.info("👈 Carga un archivo Excel en el panel izquierdo para comenzar")
    st.stop()

gestor = st.session_state.gestor

# Obtener lista de agentes para los selectores
lista_agentes = sorted(list(gestor.agentes.keys()))

# Tabs
tab1, tab2, tab3, tab4, tab5 = st.tabs(["🔄 Realizar Intercambio", "📅 Ver Cuadrante", "📋 Agentes", "📊 Resumen", "💾 Guardar"])

# Tab 1: Realizar Intercambio
with tab1:
    st.header("🔄 Realizar intercambio de turnos")
    
    col1, col2 = st.columns(2)
    
    with col1:
        agente1 = st.selectbox("👤 Selecciona agente 1", lista_agentes, key="agente1")
        if agente1:
            zona1 = gestor.agentes[agente1]["zona"]
            intercambios1 = gestor.contador_intercambios.get(agente1, 0)
            st.caption(f"Zona: {zona1} | Intercambios: {intercambios1}/{MAX_INTERCAMBIOS_POR_AGENTE}")
    
    with col2:
        agente2 = st.selectbox("👤 Selecciona agente 2", lista_agentes, key="agente2")
        if agente2:
            zona2 = gestor.agentes[agente2]["zona"]
            intercambios2 = gestor.contador_intercambios.get(agente2, 0)
            st.caption(f"Zona: {zona2} | Intercambios: {intercambios2}/{MAX_INTERCAMBIOS_POR_AGENTE}")
    
    dia = st.slider("📅 Día del mes", 1, 31, 15)
    
    if agente1 and agente2:
        turno1 = gestor.obtener_turno(agente1, dia)
        turno2 = gestor.obtener_turno(agente2, dia)
        
        col_a, col_b, col_c = st.columns(3)
        with col_a:
            st.metric(f"Turno actual de {agente1}", turno1 if turno1 else "Vacío")
        with col_b:
            st.metric(f"Turno actual de {agente2}", turno2 if turno2 else "Vacío")
        with col_c:
            dia_semana = DIAS_SEMANA_MAYO[dia-1]
            st.metric(f"Día {dia}", dia_semana)
        
        es_descanso = st.checkbox("ℹ️ Marcar como descanso", help="Activa si el agente 1 está tomando descanso")
    
    col_btn1, col_btn2 = st.columns(2)
    with col_btn1:
        if st.button("🔍 Validar intercambio", use_container_width=True):
            if agente1 == agente2:
                st.error("❌ No puedes intercambiar con el mismo agente")
            else:
                with st.spinner("Validando..."):
                    valido, errores = gestor.validar_intercambio(agente1, agente2, dia)
                    if valido:
                        st.success("✅ El intercambio es válido")
                    else:
                        st.error("❌ El intercambio no es válido")
                        for error in errores:
                            st.warning(error)
    
    with col_btn2:
        if st.button("✅ Confirmar y ejecutar intercambio", type="primary", use_container_width=True):
            if agente1 == agente2:
                st.error("❌ No puedes intercambiar con el mismo agente")
            else:
                with st.spinner("Ejecutando..."):
                    exito, mensajes = gestor.ejecutar_intercambio(agente1, agente2, dia, es_descanso)
                    if exito:
                        st.success("✅ Intercambio realizado correctamente")
                        st.balloons()
                    else:
                        for msg in mensajes:
                            st.error(msg)

# Tab 2: Ver Cuadrante
with tab2:
    st.header("📅 Ver cuadrante mensual de un agente")
    
    agente_cuadrante = st.selectbox("👤 Selecciona un agente", lista_agentes, key="cuadrante_select")
    
    if agente_cuadrante:
        if st.button("📊 Mostrar cuadrante", use_container_width=True):
            turnos = gestor.obtener_todos_turnos(agente_cuadrante)
            mostrar_cuadrante_mensual(turnos, agente_cuadrante, gestor)

# Tab 3: Agentes
with tab3:
    st.header("📋 Lista de agentes")
    
    col_f1, col_f2 = st.columns([3, 1])
    with col_f1:
        zonas = sorted(list(set(a["zona"] for a in gestor.agentes.values())))
        filtro_zona = st.selectbox("Filtrar por zona", ["Todas"] + zonas)
    with col_f2:
        busqueda = st.text_input("🔍 Buscar", placeholder="Nombre...")
    
    agentes_mostrar = gestor.agentes.items()
    if filtro_zona != "Todas":
        agentes_mostrar = [(n, i) for n, i in agentes_mostrar if i["zona"] == filtro_zona]
    if busqueda:
        agentes_mostrar = [(n, i) for n, i in agentes_mostrar if busqueda.upper() in n.upper()]
    
    st.markdown(f"**Total: {len(agentes_mostrar)} agentes**")
    
    cols = st.columns(3)
    for idx, (nombre, info) in enumerate(sorted(agentes_mostrar)):
        with cols[idx % 3]:
            with st.expander(f"{nombre[:40]}"):
                st.text(f"📍 Fila: {info['fila']}")
                st.text(f"🏢 Zona: {info['zona']}")
                st.text(f"🔄 Intercambios: {gestor.contador_intercambios.get(nombre, 0)}")

# Tab 4: Resumen
with tab4:
    st.header("📊 Resumen de intercambios")
    
    if gestor.intercambios:
        df_intercambios = pd.DataFrame(gestor.intercambios)
        st.dataframe(df_intercambios, use_container_width=True)
    else:
        st.info("No se han realizado intercambios aún")
    
    st.markdown("---")
    st.subheader("💰 Deudas pendientes")
    
    if gestor.deudas:
        for (deudor, acreedor), cantidad in gestor.deudas.items():
            st.warning(f"🔴 {deudor} debe {cantidad} día(s) a {acreedor}")
    else:
        st.success("✅ No hay deudas pendientes")
    
    st.markdown("---")
    st.subheader("📈 Intercambios por agente")
    
    if gestor.contador_intercambios:
        df_contador = pd.DataFrame([
            {"Agente": a, "Intercambios": c}
            for a, c in sorted(gestor.contador_intercambios.items(), key=lambda x: x[1], reverse=True)
        ])
        st.dataframe(df_contador, use_container_width=True)

# Tab 5: Guardar
with tab5:
    st.header("💾 Guardar cambios")
    
    st.info(f"Se han realizado {len(gestor.intercambios)} intercambio(s)")
    
    col_s1, col_s2 = st.columns(2)
    with col_s1:
        if st.button("📥 Descargar archivo modificado", type="primary", use_container_width=True):
            with st.spinner("Generando archivo..."):
                bytes_data = gestor.obtener_bytes()
                st.download_button(
                    label="⬇️ Hacer clic para descargar",
                    data=bytes_data,
                    file_name="turnos_mayo_2026_modificado.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                st.success("✅ Archivo listo para descargar")
    
    with col_s2:
        if st.button("🔄 Reiniciar sesión", use_container_width=True):
            st.session_state.gestor = None
            st.session_state.archivo_cargado = False
            st.rerun()